# backend/users/views.py

from rest_framework import generics, status, viewsets, permissions
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.decorators import action
from django.contrib.auth import authenticate
from django.http import HttpResponse
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
import pandas as pd
import io
from datetime import datetime
import threading
import time
from django.utils import timezone

from .models import User, StudentEnrollment, BulkImportJob, Profile
from .serializers import (
    UserRegisterSerializer,
    UserProfileSerializer,
    MyTokenObtainPairSerializer,
    StudentEnrollmentSerializer,
    StudentEnrollmentCreateSerializer,
    UserBasicSerializer,
    UserUpdateSerializer,
    ChildSummarySerializer
)

# The RegisterView and ProfileView remain the same
class RegisterView(generics.CreateAPIView):
    queryset = User.objects.all()
    permission_classes = (AllowAny,)
    serializer_class = UserRegisterSerializer

class ProfileView(generics.RetrieveUpdateAPIView):
    queryset = User.objects.all()
    permission_classes = (IsAuthenticated,)
    
    def get_serializer_class(self):
        # For GET (retrieve) return UserUpdateSerializer to include
        # academic, parent, and flattened profile fields expected by the frontend.
        # Keep the same serializer for updates.
        if self.request.method in ['GET', 'PUT', 'PATCH']:
            return UserUpdateSerializer
        return UserProfileSerializer

    def get_object(self):
        return self.request.user


# Add this NEW LoginView
class LoginView(APIView):
    permission_classes = [AllowAny]

    def post(self, request, *args, **kwargs):
        email = request.data.get("email")
        password = request.data.get("password")

        if not email or not password:
            return Response(
                {"error": "Please provide both email and password"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Use Django's built-in authentication
        user = authenticate(request, email=email, password=password)

        if user:
            # Check if user is using default password
            default_password = 'defaultStrongPassword25'
            if password == default_password and not user.force_password_change:
                # Mark user for forced password change
                user.force_password_change = True
                user.save()
            
            # Set user as online and update last_login
            from django.utils import timezone
            user.set_online_status(True)
            user.last_login = timezone.now()
            user.save(update_fields=['last_login'])

            # If authentication is successful, use our serializer to get tokens
            serializer = MyTokenObtainPairSerializer.get_token(user)
            refresh = str(serializer)
            access = str(serializer.access_token)

            # Use UserUpdateSerializer to get complete user data including profile
            user_serializer = UserUpdateSerializer(user)

            return Response(
                {
                    "refresh": refresh,
                    "access": access,
                    "force_password_change": user.force_password_change,
                    "user": user_serializer.data
                },
                status=status.HTTP_200_OK,
            )

        return Response(
            {"error": "Invalid Credentials"},
            status=status.HTTP_401_UNAUTHORIZED
        )


class LogoutView(APIView):
    """API endpoint for user logout"""
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        """Mark user as offline during logout"""
        try:
            user = request.user
            user.set_online_status(False)

            return Response(
                {"message": "Logout successful"},
                status=status.HTTP_200_OK
            )
        except Exception as e:
            return Response(
                {"error": f"Logout failed: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class HeartbeatView(APIView):
    """API endpoint for tracking user activity"""
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        """Update user's last seen timestamp and ensure they're marked as online"""
        try:
            user = request.user
            user.update_last_seen()

            # Ensure user is marked as online (in case they weren't)
            if not user.is_online:
                user.set_online_status(True)

            return Response(
                {
                    "status": "success",
                    "last_seen": user.last_seen,
                    "is_online": user.is_online
                },
                status=status.HTTP_200_OK
            )
        except Exception as e:
            return Response(
                {"error": f"Heartbeat failed: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class CleanupInactiveUsersView(APIView):
    """API endpoint for cleanup of inactive users (admin only)"""
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        """Mark users as offline if they haven't been active for a while"""
        if not request.user.role == 'ADMIN':
            return Response(
                {"error": "Admin access required"},
                status=status.HTTP_403_FORBIDDEN
            )

        try:
            from datetime import timedelta

            # Mark users as offline if they haven't been seen in 10 minutes
            inactive_threshold = timezone.now() - timedelta(minutes=10)

            updated_count = User.objects.filter(
                is_online=True,
                last_seen__lt=inactive_threshold
            ).update(is_online=False)

            return Response(
                {
                    "status": "success",
                    "users_marked_offline": updated_count,
                    "threshold_minutes": 10
                },
                status=status.HTTP_200_OK
            )
        except Exception as e:
            return Response(
                {"error": f"Cleanup failed: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ChangePasswordView(APIView):
    """API endpoint for changing user password"""
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        user = request.user
        current_password = request.data.get('current_password')
        new_password = request.data.get('new_password')
        confirm_password = request.data.get('confirm_password')

        if not all([current_password, new_password, confirm_password]):
            return Response(
                {"error": "All fields are required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Check if new passwords match
        if new_password != confirm_password:
            return Response(
                {"error": "New passwords do not match"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validate current password
        if not user.check_password(current_password):
            return Response(
                {"error": "Current password is incorrect"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validate new password strength
        if len(new_password) < 8:
            return Response(
                {"error": "Password must be at least 8 characters long"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Check if new password is different from current
        if user.check_password(new_password):
            return Response(
                {"error": "New password must be different from current password"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Set new password and clear force password change flag
        user.set_password(new_password)
        user.force_password_change = False
        user.save()

        return Response(
            {"message": "Password changed successfully"},
            status=status.HTTP_200_OK
        )


# =====================================
# PERMISSION CLASSES
# =====================================

class IsTeacherOrAdmin(permissions.BasePermission):
    """Permission for teachers and admins"""
    def has_permission(self, request, view):
        return request.user.is_authenticated and request.user.role in ['TEACHER', 'ADMIN']


class IsAdminOrReadOnly(permissions.BasePermission):
    """Permission for admins to modify, others to read only"""
    def has_permission(self, request, view):
        if request.method in permissions.SAFE_METHODS:
            return request.user.is_authenticated
        return request.user.is_authenticated and request.user.role == 'ADMIN'


# =====================================
# USER VIEWSET
# =====================================

class UserViewSet(viewsets.ModelViewSet):
    """ViewSet for listing, retrieving, and updating users"""
    permission_classes = [IsAdminOrReadOnly]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['is_active', 'parent', 'profile__school_subject', 'profile__teachable_grades']
    search_fields = [
        'first_name',
        'last_name',
        'email',
        'profile__phone',
        'children__first_name',
        'children__last_name'
    ]
    ordering_fields = ['first_name', 'last_name', 'email', 'created_at']
    ordering = ['last_name', 'first_name']

    def get_serializer_class(self):
        if self.action in ['update', 'partial_update']:
            return UserUpdateSerializer
        elif self.action == 'retrieve':
            return UserUpdateSerializer  # Use same serializer for retrieve to get profile data
        return UserBasicSerializer
    
    def get_queryset(self):
        # Base queryset with optimized relations
        queryset = User.objects.select_related(
            'profile',  # User profile data
            'parent',   # Parent information for students
            'parent__profile',  # Parent profile data
            'profile__school_subject'  # Subject specialization for teachers
        ).prefetch_related(
            'children',
            'student_enrollments__school_class__grade__educational_level',  # Academic information
            'student_enrollments__academic_year',  # Academic year information
            'profile__teachable_grades',  # Teachable grades for teachers
            'teaching_classes__grade__educational_level',  # Teacher's classes
            'teaching_classes__academic_year'  # Academic year for teacher's classes
        ).all()

        # Additional filtering by role (supports comma-separated roles)
        role = self.request.query_params.get('role')
        if role:
            roles = [r.strip().upper() for r in role.split(',')]
            if len(roles) > 1:
                queryset = queryset.filter(role__in=roles)
            else:
                queryset = queryset.filter(role=roles[0])

        # Filter by profile position
        position = self.request.query_params.get('position')
        if position:
            queryset = queryset.filter(profile__position=position.upper())

        # Filter teachers by subject specialization
        subject_id = self.request.query_params.get('subject_id')
        if subject_id and role and role.upper() == 'TEACHER':
            queryset = queryset.filter(profile__school_subject_id=subject_id)

        return queryset.distinct()

    @action(detail=False, methods=['get'], url_path='available-drivers')
    def available_drivers(self, request):
        """Return minimal info about active staff members whose profile position is Driver."""
        include_inactive = request.query_params.get('include_inactive', 'false').lower() in ['true', '1', 'yes']
        drivers = self.get_queryset().filter(profile__position=Profile.Position.DRIVER)
        if not include_inactive:
            drivers = drivers.filter(is_active=True)

        drivers = drivers.order_by('first_name', 'last_name', 'email')
        data = [
            {
                'id': user.id,
                'full_name': user.full_name,
                'email': user.email,
                'position': user.profile.position,
            }
            for user in drivers
        ]
        return Response(data)

    @action(detail=False, methods=['get'])
    def my_classes(self, request):
        """Get current teacher's assigned classes"""
        if request.user.role != 'TEACHER':
            return Response({'error': 'Only teachers can access this endpoint'}, status=403)

        from schools.models import SchoolClass
        from schools.serializers import SchoolClassSerializer

        # Get teacher's assigned classes with related data
        classes = SchoolClass.objects.filter(
            teachers=request.user
        ).select_related(
            'grade',
            'grade__educational_level',
            'track',
            'academic_year'
        ).prefetch_related(
            'student_enrollments__student'
        )

        serializer = SchoolClassSerializer(classes, many=True)
        return Response({
            'classes': serializer.data,
            'teacher': {
                'id': request.user.id,
                'name': request.user.full_name,
                'subject': request.user.profile.school_subject.name if request.user.profile.school_subject else None
            }
        })

    @action(detail=False, methods=['get'])
    def my_students(self, request):
        """Get all students enrolled in classes taught by the current teacher"""
        if request.user.role != 'TEACHER':
            return Response({'error': 'Only teachers can access this endpoint'}, status=403)

        from schools.models import SchoolClass
        from attendance.models import TimetableSession
        
        # 1. Get classes taught by this teacher via Timetable
        class_ids = TimetableSession.objects.filter(
            teacher=request.user,
            is_active=True
        ).values_list('timetable__school_class_id', flat=True).distinct()
        
        classes = SchoolClass.objects.filter(id__in=class_ids)

        # 2. Get active enrollments for these classes
        enrollments = StudentEnrollment.objects.filter(
            school_class__in=classes,
            is_active=True
        ).select_related(
            'student', 
            'student__profile',
            'school_class',
            'school_class__grade',
            'school_class__grade__educational_level'
        ).order_by('student__last_name', 'student__first_name')

        # 3. Process students (handle duplicates if student is in multiple classes of the teacher)
        # We want to list students, and maybe show which class they are in.
        # Since a student might be in multiple classes taught by the same teacher (unlikely in primary/secondary but possible),
        # we'll return a list of student objects, each containing a list of classes they are enrolled in with this teacher.
        
        students_map = {}
        
        for enrollment in enrollments:
            student_id = enrollment.student.id
            if student_id not in students_map:
                student = enrollment.student
                students_map[student_id] = {
                    'id': student.id,
                    'full_name': student.full_name,
                    'email': student.email,
                    'avatar': student.profile.profile_picture_url if hasattr(student, 'profile') else None,
                    'gender': student.profile.gender if hasattr(student, 'profile') else None,
                    'phone': student.profile.phone if hasattr(student, 'profile') else None,
                    'classes': []
                }
            
            # Add class info
            students_map[student_id]['classes'].append({
                'id': enrollment.school_class.id,
                'name': enrollment.school_class.name,
                'grade': enrollment.school_class.grade.name
            })
        
        return Response({
            'students': list(students_map.values()),
            'count': len(students_map)
        })

    @action(detail=False, methods=['get'])
    def my_teachable_classes(self, request):
        """Get all classes that match teacher's subject and teachable grades (for assignment purposes)"""
        if request.user.role != 'TEACHER':
            return Response({'error': 'Only teachers can access this endpoint'}, status=403)

        from schools.models import SchoolClass
        from schools.serializers import SchoolClassSerializer

        # Get classes that match teacher's teachable grades and current academic year
        teachable_grades = request.user.profile.teachable_grades.all()
        if not teachable_grades.exists():
            return Response({
                'classes': [],
                'message': 'No teachable grades assigned to teacher'
            })

        # Get current academic year
        from schools.models import AcademicYear
        current_year = AcademicYear.objects.filter(is_current=True).first()
        if not current_year:
            return Response({
                'classes': [],
                'message': 'No current academic year set'
            })

        classes = SchoolClass.objects.filter(
            grade__in=teachable_grades,
            academic_year=current_year
        ).select_related(
            'grade',
            'grade__educational_level',
            'track',
            'academic_year'
        ).prefetch_related(
            'teachers',
            'student_enrollments__student'
        )

        serializer = SchoolClassSerializer(classes, many=True)
        return Response({
            'classes': serializer.data,
            'teacher': {
                'id': request.user.id,
                'name': request.user.full_name,
                'subject': request.user.profile.school_subject.name if request.user.profile.school_subject else None,
                'teachable_grades': [{'id': g.id, 'name': g.name} for g in teachable_grades]
            }
        })

    @action(detail=False, methods=['get'], url_path='staff-position-options')
    def staff_position_options(self, request):
        """Return available staff position choices with multilingual labels."""
        language = 'en'
        lang_param = request.query_params.get('lang') if hasattr(request, 'query_params') else None
        header_lang = request.headers.get('Accept-Language') if hasattr(request, 'headers') else None
        language = (lang_param or getattr(request, 'LANGUAGE_CODE', None) or language)
        if header_lang and not lang_param:
            language = header_lang.split(',')[0]
        language = (language or 'en').split('-')[0]

        options = []
        for value, _ in Profile.Position.choices:
            labels = Profile.POSITION_LABELS.get(value, {})
            options.append({
                'value': value,
                'label': labels.get(language, labels.get('en')),
                'labels': labels
            })

        return Response({'positions': options})

    @action(detail=True, methods=['get'], url_path='children')
    def children(self, request, pk=None):
        """Get all children (students) of a parent"""
        parent = self.get_object()

        if parent.role != User.Role.PARENT:
            return Response(
                {'error': 'User is not a parent'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Get all children with optimized queries
        children = parent.children.select_related(
            'profile'
        ).prefetch_related(
            'student_enrollments__school_class__grade',
            'student_enrollments__academic_year'
        ).filter(is_active=True)

        # Performance optimization: Include counts for pending homework and absences
        from django.db.models import Count, Q, OuterRef, Subquery
        from homework.models import Homework, Submission
        from attendance.models import StudentAbsenceFlag

        # We can't easily annotate HW counts with simple Count because it depends on class and student
        # So we'll iterate for now or use a more complex subquery if performance issues arise.
        # Given this is likely a small number of children, a clean loop or SerializerMethodField is safer.
        
        serializer = ChildSummarySerializer(children, many=True, context={'request': request})

        return Response({
            'parent': {
                'id': parent.id,
                'name': parent.full_name,
                'email': parent.email,
                'phone': parent.profile.phone if hasattr(parent, 'profile') else None
            },
            'children': serializer.data,
            'total_children': children.count()
        })

    @action(detail=True, methods=['get'], url_path='teachers')
    def teachers(self, request, pk=None):
        """Get all teachers for a specific student"""
        student = self.get_object()

        if student.role != User.Role.STUDENT:
            return Response(
                {'error': 'User is not a student'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Get the student's active enrollment
        enrollment = student.student_enrollments.filter(is_active=True).select_related('school_class').first()
        if not enrollment or not enrollment.school_class:
            return Response({'teachers': []})

        # Reuse SchoolClassSerializer logic to get teachers
        from schools.serializers import SchoolClassSerializer
        # We only need the teachers field from the serializer
        class_serializer = SchoolClassSerializer(enrollment.school_class, context={'request': request})
        teachers = class_serializer.data.get('teachers', [])

        return Response({'teachers': teachers})


# =====================================
# STUDENT ENROLLMENT VIEWSET
# =====================================

class StudentEnrollmentViewSet(viewsets.ModelViewSet):
    """ViewSet for student enrollments"""
    queryset = StudentEnrollment.objects.select_related(
        'student', 'school_class', 'academic_year', 'school_class__grade', 'school_class__grade__educational_level'
    ).all()
    permission_classes = [IsTeacherOrAdmin]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = {
        'school_class__grade__educational_level': ['exact'],
        'school_class__grade': ['exact'],
        'school_class': ['exact'],
        'academic_year': ['exact'],
        'is_active': ['exact'],
    }
    search_fields = ['student__first_name', 'student__last_name', 'student__email', 'student_number']
    ordering_fields = ['student__last_name', 'student__first_name', 'enrollment_date']
    ordering = ['student__last_name', 'student__first_name']

    def get_serializer_class(self):
        if self.action == 'create':
            return StudentEnrollmentCreateSerializer
        return StudentEnrollmentSerializer

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """
        Return aggregated statistics for student enrollments.
        Returns total count, active/inactive counts, gender distribution, and level distribution.
        """
        from django.db.models import Count, Q
        
        # Get base queryset (filtered by permissions implicitly via get_queryset not fully applied here, 
        # but we should respect general filtering if needed. For now, stats are global or per school context)
        # Ideally, we should filter by the user's scope if multi-tenant. 
        # Assuming current implementation expects all students visible to admin/teacher.
        
        # Get base queryset
        queryset = self.filter_queryset(self.get_queryset())
        
        # Clear ordering to ensure proper grouping for ALL aggregations
        queryset = queryset.order_by()
        
        # Total and Active/Inactive
        total_students = queryset.count()
        active_students = queryset.filter(is_active=True).count()
        inactive_students = total_students - active_students
        
        # Gender Distribution (using student__profile__gender)
        # Note: Gender is stored in Profile model linked to User (student)
        gender_stats = queryset.values('student__profile__gender').annotate(count=Count('id'))
        males = 0
        females = 0
        for item in gender_stats:
            gender = item.get('student__profile__gender')
            if gender == 'MALE':
                males += item['count']
            elif gender == 'FEMALE':
                females += item['count']
        
        # Calculate unknown gender count to ensure numbers match up
        unknown_gender = total_students - males - females
        
        # Level Distribution
        # Group by educational level
        level_stats = queryset.values(
            'school_class__grade__educational_level__id',
            'school_class__grade__educational_level__name',
            'school_class__grade__educational_level__name_arabic',
            'school_class__grade__educational_level__name_french'
        ).annotate(count=Count('id'))
        
        level_counts = []
        for item in level_stats:
            if item['school_class__grade__educational_level__id']:
                level_counts.append({
                    'id': item['school_class__grade__educational_level__id'],
                    'name': item['school_class__grade__educational_level__name'],
                    'name_arabic': item['school_class__grade__educational_level__name_arabic'],
                    'name_french': item['school_class__grade__educational_level__name_french'],
                    'count': item['count']
                })

        # Grade Distribution
        grade_stats = queryset.values(
            'school_class__grade__id',
            'school_class__grade__name',
            'school_class__grade__name_arabic',
            'school_class__grade__name_french'
        ).annotate(count=Count('id')).order_by('school_class__grade__educational_level__order', 'school_class__grade__grade_number')
        
        grade_counts = []
        for item in grade_stats:
            if item['school_class__grade__id']:
                grade_counts.append({
                    'id': item['school_class__grade__id'],
                    'name': item['school_class__grade__name'],
                    'name_arabic': item['school_class__grade__name_arabic'],
                    'name_french': item['school_class__grade__name_french'],
                    'count': item['count']
                })
        
        return Response({
            'total': total_students,
            'active': active_students,
            'inactive': inactive_students,
            'males': males,
            'females': females,
            'unknown': unknown_gender,
            'levelCounts': level_counts,
            'gradeCounts': grade_counts
        })


# =====================================
# BULK IMPORT VIEWS
# =====================================

class StudentBulkImportView(APIView):
    """
    View for bulk import of students via Excel template
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        """Download Excel template for bulk student import"""
        try:
            # Get educational structure parameters (if provided)
            level_id = request.GET.get('level_id')
            grade_id = request.GET.get('grade_id')  
            class_id = request.GET.get('class_id')
            academic_year_id = request.GET.get('academic_year_id')
            
            # Get educational structure names for template info
            level_name = ''
            grade_name = ''
            class_name = ''
            academic_year = ''
            
            if level_id and grade_id and class_id and academic_year_id:
                from schools.models import EducationalLevel, Grade, SchoolClass, AcademicYear
                try:
                    level = EducationalLevel.objects.get(id=level_id)
                    grade = Grade.objects.get(id=grade_id)
                    school_class = SchoolClass.objects.get(id=class_id)
                    academic_year_obj = AcademicYear.objects.get(id=academic_year_id)
                    
                    level_name = level.name
                    grade_name = grade.name
                    class_name = school_class.name
                    academic_year = academic_year_obj.year
                except:
                    pass
            
            # Create Excel file
            import openpyxl
            buffer = io.BytesIO()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Students"
            
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            
            # Styles
            header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            header_font = Font(bold=True, size=11)
            centered = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Helper to set border
            def set_border(cell):
                cell.border = thin_border
                cell.alignment = centered

            # 1. Add Note in Row 3 (Merged cells, Right Aligned as per screenshot)
            # Alignment seems to be above the last few columns. With shift, Note is at N.
            ws.merge_cells('J3:M3')
            ws['J3'] = "ملاحظة : الجنس : M / F"
            ws['J3'].alignment = Alignment(horizontal="right", vertical="center")
            ws['J3'].font = Font(bold=True)

            ws.merge_cells('G4:M4')
            ws['G4'] = "قم بحذف معطيات السطر الأول (المثال) وقم بملء البيانات بالطريقة التالية :"
            ws['G4'].alignment = Alignment(horizontal="right", vertical="center")
            ws['G4'].font = Font(bold=True)

            # 2. Define Headers
            # Row 6: Arabic Headers
            arabic_headers = [
                'الاسم الشخصي بالفرنسية',    # B: First Name French
                'الاسم العائلي بالفرنسية',    # C: Last Name French
                'الاسم الشخصي بالعربية',      # D: First Name Arabic
                'الاسم العائلي بالعربية',      # E: Last Name Arabic
                'الجنس',                      # F: Gender
                'تاريخ الازدياد',             # G: DOB
                'رقم الهاتف',                 # H: Phone
                'العنوان',                    # I: Address
                'اسم ولي الامر',              # J: Parent First Name
                'الاسم العائلي لولي الامر',   # K: Parent Last Name
                'هاتف ولي الامر',             # L: Parent Phone
                'رقم الطوارئ',                # M: Emergency Phone
                'ملاحظات'                     # N: Notes
            ]
            
            # Row 7: French/English Headers (matching screenshot IDs)
            french_headers = [
                'Prenom',
                'Nom',
                'Arabic prenom',
                'Arabic nom',
                'Genre',
                'Date de naissance',
                'Phone',
                'Adress',
                'Prenom du Parent',
                'Nom du parent',
                'Phone du parent',
                'Emegency Phone',
                'Note'
            ]
            
            # Write Row 6 (Arabic) - Start at Column 2
            for col_idx, text in enumerate(arabic_headers, 2):
                cell = ws.cell(row=6, column=col_idx, value=text)
                cell.fill = header_fill
                cell.font = header_font
                set_border(cell)
                
            # Write Row 7 (French) - Start at Column 2
            for col_idx, text in enumerate(french_headers, 2):
                cell = ws.cell(row=7, column=col_idx, value=text)
                cell.fill = header_fill
                cell.font = header_font
                set_border(cell)

            # 3. Sample Data (Row 8) - Start at Column 2
            sample_row = [
                'Younes', 'El bettate', 'يونس', 'البتات', 'M', '1992-05-15', '0600000000', '123 Main St', 
                'Mohamed', 'El bettate', '0611111111', '0622222222', 'Bon eleve'
            ]
            for col_idx, value in enumerate(sample_row, 2):
                cell = ws.cell(row=8, column=col_idx, value=value)
                set_border(cell)

            # Set column widths
            # A is empty/narrow. B-N match the data.
            # Adjusted widths to "fit screen" better
            column_widths_map = {
                1: 8,   # A: Empty
                2: 13,  # B: Prenom
                3: 13,  # C: Nom
                4: 13,  # D: Ar prenom
                5: 13,  # E: Ar nom
                6: 13,  # F: Genre
                7: 13,  # G: Date
                8: 13,  # H: Phone
                9: 25,  # I: Adress (Wider)
                10: 13, # J: Parent First
                11: 13, # K: Parent Last
                12: 13, # L: Parent Phone
                13: 13, # M: Emergency Phone
                14: 25  # N: Note
            }
            
            for col_idx, width in column_widths_map.items():
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = width

            # Add Metadata/Instructions Sheet (hidden or secondary)
            ws_meta = wb.create_sheet("Info")
            ws_meta.append(['Property', 'Value'])
            ws_meta.append(['Level', level_name])
            ws_meta.append(['Grade', grade_name])
            ws_meta.append(['Class', class_name])
            ws_meta.append(['Academic Year', academic_year])
            
            wb.save(buffer)
            buffer.seek(0)
            
            # Create response
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="student_import_template_{datetime.now().strftime("%Y%m%d")}.xlsx'
            
            return response
            
        except Exception as e:
            return Response(
                {'error': f'Failed to generate template: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def post(self, request):
        """Process uploaded Excel file for bulk student import"""
        try:
            # Debug: Log the incoming request
            import logging
            logger = logging.getLogger(__name__)
            
            # Check if file was uploaded
            if 'file' not in request.FILES:
                return Response(
                    {'error': 'No file uploaded'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Get educational structure parameters
            level_id = request.data.get('level_id')
            grade_id = request.data.get('grade_id')
            class_id = request.data.get('class_id')
            academic_year_id = request.data.get('academic_year_id')
            
            # Validate educational structure parameters
            if not all([level_id, grade_id, class_id, academic_year_id]):
                return Response(
                    {'error': 'Educational structure parameters (level_id, grade_id, class_id, academic_year_id) are required'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            uploaded_file = request.FILES['file']
            
            # Validate file type
            if not uploaded_file.name.endswith(('.xlsx', '.xls')):
                return Response(
                    {'error': 'Please upload an Excel file (.xlsx or .xls)'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Read Excel file
            try:
                df = pd.read_excel(uploaded_file, sheet_name='Students', header=6)
                # Filter out completely empty rows
                df = df.dropna(how='all')
                # Filter out rows where the mandatory fields are empty (simple check)
                if not df.empty:
                     df = df[df['Prenom'].notna()]
            except Exception as e:
                return Response({'error': f'Failed to read Excel file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Validate required columns (simplified - no educational structure columns)
            required_columns = [
                'Prenom', 'Nom', 
                'Arabic prenom', 'Arabic nom'
            ]
            
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                return Response(
                    {'error': f'Missing required columns: {", ".join(missing_columns)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Process and validate data with pre-selected educational structure
            preview_mode = request.data.get('preview', 'true').lower() == 'true'
            educational_structure = {
                'level_id': int(level_id),
                'grade_id': int(grade_id), 
                'class_id': int(class_id),
                'academic_year_id': int(academic_year_id)
            }
            
            if preview_mode:
                # For preview, process synchronously
                results = self._process_student_data(df, preview_mode, educational_structure)
                return Response(results, status=status.HTTP_200_OK)
            else:
                # For actual import, process asynchronously
                job = BulkImportJob.objects.create(
                    created_by=request.user,
                    total_records=len(df),
                    current_status='Initializing import...'
                )
                
                # Start background processing
                thread = threading.Thread(
                    target=self._process_import_async,
                    args=(job.job_id, df, educational_structure)
                )
                thread.daemon = True
                thread.start()
                
                return Response({
                    'job_id': str(job.job_id),
                    'status': 'started',
                    'message': 'Import job started successfully'
                }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response(
                {'error': f'Import failed: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def _process_import_async(self, job_id, df, educational_structure):
        """Process import in background thread"""
        import logging
        logger = logging.getLogger(__name__)
        
        try:
            job = BulkImportJob.objects.get(job_id=job_id)
            job.status = BulkImportJob.Status.PROCESSING
            job.started_at = timezone.now()
            job.save()
            
            # Process the import with progress tracking
            results = self._process_student_data_with_progress(df, job, educational_structure)
            
            # Mark job as completed
            job.mark_completed(results)
            logger.info(f"Import job {job_id} completed successfully")
            
        except Exception as e:
            logger.error(f"Import job {job_id} failed: {str(e)}")
            try:
                job = BulkImportJob.objects.get(job_id=job_id)
                job.mark_failed(str(e))
            except:
                pass

    def _generate_unique_email_for_import(self, first_name, last_name, role, used_emails):
        """
        Generates a unique email for the import process.
        Checks against the database and a set of emails used in the current batch.
        """
        import re
        from .models import User

        # Clean names and get initial
        clean_last_name = re.sub(r'[^a-z0-9]', '', last_name.lower().replace(' ', '')).strip()
        initial = first_name[0].lower() if first_name else 'u'
        
        # Determine email domain
        school_name = 'madrasti' # This could be a setting
        role_suffix_map = {'STUDENT': 'students', 'PARENT': 'parents', 'TEACHER': 'teachers', 'ADMIN': 'team', 'STAFF': 'team'}
        domain_suffix = role_suffix_map.get(role, 'users')
        email_domain = f"@{school_name}-{domain_suffix}.com"
        
        # Base email part (e.g., j.doe)
        base_email_part = f"{initial}.{clean_last_name}"
        
        # First attempt - clean email without numbers
        candidate_email = base_email_part + email_domain
        
        # Check for uniqueness and add simple incremental number if needed
        counter = 2  # Start with 2 for the second attempt (first attempt has no number)
        while User.objects.filter(email=candidate_email).exists() or candidate_email in used_emails:
            candidate_email = f"{base_email_part}{counter}{email_domain}"
            counter += 1
            # Safety break to avoid infinite loop
            if counter > 1000:
                break
            
        return candidate_email

    def _process_student_data(self, df, preview_mode=True, educational_structure=None):
        """Process student data from DataFrame"""
        results = {
            'total_rows': len(df),
            'processed_rows': 0,
            'successful_imports': 0,
            'errors': [],
            'warnings': [],
            'preview_data': [] if preview_mode else None,
            'created_students': [] if not preview_mode else None,
            'created_parents': [] if not preview_mode else None
        }
        
        generated_emails_in_batch = set()
        
        for index, row in df.iterrows():
            try:
                # Skip empty rows
                if pd.isna(row.get('Prenom')) or pd.isna(row.get('Nom')):
                    continue
                
                results['processed_rows'] += 1
                row_number = index + 2  # Excel row number (accounting for header)

                # --- Basic Row Validation ---
                row_errors = []
                # Map Excel columns to internal variable names
                col_map = {
                    'first_name': 'Prenom',
                    'last_name': 'Nom',
                    'ar_first_name': 'Arabic prenom',
                    'ar_last_name': 'Arabic nom',
                    'gender': 'Genre',
                    'dob': 'Date de naissance',
                    'phone': 'Phone',
                    'address': 'Adress', # Typo "Adress" in template to match screenshot
                    'parent_first': 'Prenom du Parent',
                    'parent_last': 'Nom du parent',
                    'parent_phone': 'Phone du parent',
                    'emergency_phone': 'Emegency Phone', # Typo matches template
                    'note': 'Note'
                }

                required_fields = {
                    col_map['first_name']: 'First name', 
                    col_map['last_name']: 'Last name', 
                    col_map['ar_first_name']: 'Arabic first name', 
                    col_map['ar_last_name']: 'Arabic last name'
                }
                
                for field, display_name in required_fields.items():
                    if pd.isna(row.get(field)) or str(row.get(field)).strip() == '':
                        row_errors.append(f'{display_name} is required.')
                
                if not pd.isna(row.get(col_map['dob'])):
                    try: pd.to_datetime(row[col_map['dob']])
                    except: row_errors.append('Invalid date format. Use YYYY-MM-DD.')
                
                if row_errors:
                    results['errors'].append({'row': row_number, 'error': ' '.join(row_errors)})
                    continue
                # --- End Basic Validation ---

                # --- Prepare Student Data ---
                from datetime import date
                student_first_name = str(row[col_map['first_name']]).strip()
                student_last_name = str(row[col_map['last_name']]).strip()
                
                # Generate final, unique email before validation
                student_email = self._generate_unique_email_for_import(
                    student_first_name, student_last_name, 'STUDENT', generated_emails_in_batch
                )

                # Parse Gender
                gender_raw = str(row.get(col_map['gender'], '')).strip().upper()
                gender_val = None
                if gender_raw in ['M', 'MALE', 'HOMME']:
                    gender_val = 'MALE'
                elif gender_raw in ['F', 'FEMALE', 'FEMME']:
                    gender_val = 'FEMALE'

                student_data = {
                    'email': student_email,
                    'password': 'defaultStrongPassword25',
                    'first_name': student_first_name,
                    'last_name': student_last_name,
                    'role': 'STUDENT',
                    'ar_first_name': str(row[col_map['ar_first_name']]).strip(),
                    'ar_last_name': str(row[col_map['ar_last_name']]).strip(),
                    'enrollment_date': date.today(),
                    'gender': gender_val
                }
                
                if educational_structure:
                    student_data['school_class_id'] = educational_structure['class_id']
                    student_data['academic_year_id'] = educational_structure['academic_year_id']
                
                # Map optional fields
                optional_mapping = {
                    'phone': col_map['phone'], 
                    'address': col_map['address'], 
                    'bio': col_map['note'],
                    'emergency_contact_phone': col_map['emergency_phone'],
                    'parent_first_name': col_map['parent_first'], 
                    'parent_last_name': col_map['parent_last'], 
                    'parent_phone': col_map['parent_phone']
                }
                
                for field, column in optional_mapping.items():
                    if column in row and not pd.isna(row[column]):
                        student_data[field] = str(row[column]).strip()
                
                if col_map['dob'] in row and not pd.isna(row[col_map['dob']]):
                    student_data['date_of_birth'] = pd.to_datetime(row[col_map['dob']]).date()
                # --- End Data Preparation ---

                if preview_mode:
                    generated_emails_in_batch.add(student_email) # Add to set even in preview
                    
                    # Generate predicted parent email if parent data exists
                    predicted_parent_email = None
                    if student_data.get('parent_first_name') and student_data.get('parent_last_name'):
                        predicted_parent_email = self._generate_unique_email_for_import(
                            student_data['parent_first_name'], 
                            student_data['parent_last_name'], 
                            'PARENT', 
                            generated_emails_in_batch
                        )
                    
                    preview_item = {
                        'row_number': row_number,
                        'student_name': f"{student_first_name} {student_last_name}",
                        'arabic_name': f"{row[col_map['ar_first_name']]} {row[col_map['ar_last_name']]}",
                        'parent_name': f"{row.get(col_map['parent_first'], '')} {row.get(col_map['parent_last'], '')}".strip(),
                        'predicted_student_email': student_email,
                        'predicted_parent_email': predicted_parent_email or 'No parent data provided'
                    }
                    if results['preview_data'] is None: results['preview_data'] = []
                    results['preview_data'].append(preview_item)
                else:
                    # Final Import: Validate and Save
                    serializer = UserRegisterSerializer(data=student_data)
                    if serializer.is_valid():
                        try:
                            student = serializer.save()
                            generated_emails_in_batch.add(student.email) # Add final email to set
                            results['successful_imports'] += 1
                            results['created_students'].append({'id': student.id, 'email': student.email, 'full_name': student.full_name, 'row_number': row_number})
                            
                            # Track parent creation for results
                            if student_data.get('parent_first_name') and student.parent:
                                # Check if this parent was already tracked in this batch
                                parent_already_tracked = any(
                                    p['id'] == student.parent.id 
                                    for p in results['created_parents']
                                )
                                if not parent_already_tracked:
                                    results['created_parents'].append({
                                        'id': student.parent.id, 
                                        'email': student.parent.email, 
                                        'full_name': student.parent.full_name,
                                        'children_count': student.parent.children.count()
                                    })

                        except Exception as save_error:
                            results['errors'].append({'row': row_number, 'error': f"Save Error: {save_error}"})
                    else:
                        results['errors'].append({'row': row_number, 'error': f"Validation failed: {serializer.errors}"})
                                
            except Exception as e:
                results['errors'].append({'row': index + 2, 'error': str(e)})
        
        return results

    def _process_student_data_with_progress(self, df, job, educational_structure):
        """Process student data with progress tracking"""
        results = {
            'total_rows': len(df),
            'processed_rows': 0,
            'successful_imports': 0,
            'errors': [],
            'warnings': [],
            'created_students': [],
            'created_parents': []
        }
        
        generated_emails_in_batch = set()
        total_rows = len(df)
        
        job.update_progress(5, "Validating data...")
        
        for index, row in df.iterrows():
            try:
                # Skip empty rows
                if pd.isna(row.get('Prenom')) or pd.isna(row.get('Nom')):
                    continue
                
                results['processed_rows'] += 1
                row_number = index + 2  # Excel row number (accounting for header)
                
                # Update progress after each record for smoother progress bar
                progress = 10 + (results['processed_rows'] / total_rows * 85)  # 10% to 95%
                status = f"Processing student {results['processed_rows']} of {total_rows}..."
                job.update_progress(int(progress), status)

                # --- Basic Row Validation ---
                row_errors = []
                # Map Excel columns to internal variable names
                # Updated to match the new 2-row header format where we read the French row (Row 7)
                col_map = {
                    'first_name': 'Prenom',
                    'last_name': 'Nom',
                    'ar_first_name': 'Arabic prenom',
                    'ar_last_name': 'Arabic nom',
                    'gender': 'Genre',
                    'dob': 'Date de naissance',
                    'phone': 'Phone',
                    'address': 'Adress', # Typo "Adress" in template to match screenshot
                    'parent_first': 'Prenom du Parent',
                    'parent_last': 'Nom du parent',
                    'parent_phone': 'Phone du parent',
                    'emergency_phone': 'Emegency Phone', # Typo matches template
                    'note': 'Note'
                }

                required_fields = {
                    col_map['first_name']: 'First name', 
                    col_map['last_name']: 'Last name', 
                    col_map['ar_first_name']: 'Arabic first name', 
                    col_map['ar_last_name']: 'Arabic last name'
                }
                
                for field, display_name in required_fields.items():
                    if pd.isna(row.get(field)) or str(row.get(field)).strip() == '':
                        row_errors.append(f'{display_name} is required.')
                
                if not pd.isna(row.get(col_map['dob'])):
                    try: pd.to_datetime(row[col_map['dob']])
                    except: row_errors.append('Invalid date format. Use YYYY-MM-DD.')
                
                if row_errors:
                    results['errors'].append({'row': row_number, 'error': ' '.join(row_errors)})
                    continue
                # --- End Basic Validation ---
                
                # --- Prepare Student Data ---
                from datetime import date
                student_first_name = str(row[col_map['first_name']]).strip()
                student_last_name = str(row[col_map['last_name']]).strip()
                
                # Generate final, unique email before validation
                student_email = self._generate_unique_email_for_import(
                    student_first_name, student_last_name, 'STUDENT', generated_emails_in_batch
                )

                # Parse Gender
                gender_raw = str(row.get(col_map['gender'], '')).strip().upper()
                gender_val = None
                if gender_raw in ['M', 'MALE', 'HOMME']:
                    gender_val = 'MALE'
                elif gender_raw in ['F', 'FEMALE', 'FEMME']:
                    gender_val = 'FEMALE'

                student_data = {
                    'email': student_email,
                    'password': 'defaultStrongPassword25',
                    'first_name': student_first_name,
                    'last_name': student_last_name,
                    'role': 'STUDENT',
                    'ar_first_name': str(row[col_map['ar_first_name']]).strip(),
                    'ar_last_name': str(row[col_map['ar_last_name']]).strip(),
                    'enrollment_date': date.today(),
                    'gender': gender_val
                }
                
                if educational_structure:
                    student_data['school_class_id'] = educational_structure['class_id']
                    student_data['academic_year_id'] = educational_structure['academic_year_id']
                
                # Map optional fields
                optional_mapping = {
                    'phone': col_map['phone'], 
                    'address': col_map['address'], 
                    'bio': col_map['note'],
                    'emergency_contact_phone': col_map['emergency_phone'],
                    'parent_first_name': col_map['parent_first'], 
                    'parent_last_name': col_map['parent_last'], 
                    'parent_phone': col_map['parent_phone']
                }
                
                for field, column in optional_mapping.items():
                    if column in row and not pd.isna(row[column]):
                        student_data[field] = str(row[column]).strip()
                
                if col_map['dob'] in row and not pd.isna(row[col_map['dob']]):
                    student_data['date_of_birth'] = pd.to_datetime(row[col_map['dob']]).date()
                # --- End Data Preparation ---

                # Final Import: Validate and Save
                serializer = UserRegisterSerializer(data=student_data)
                if serializer.is_valid():
                    try:
                        student = serializer.save()
                        generated_emails_in_batch.add(student.email) # Add final email to set
                        results['successful_imports'] += 1
                        results['created_students'].append({
                            'id': student.id, 
                            'email': student.email, 
                            'full_name': student.full_name, 
                            'row_number': row_number
                        })
                        
                        # Track parent creation for results
                        if student_data.get('parent_first_name') and student.parent:
                            # Check if this parent was already tracked in this batch
                            parent_already_tracked = any(
                                p['id'] == student.parent.id 
                                for p in results['created_parents']
                            )
                            if not parent_already_tracked:
                                results['created_parents'].append({
                                    'id': student.parent.id, 
                                    'email': student.parent.email, 
                                    'full_name': student.parent.full_name,
                                    'children_count': student.parent.children.count()
                                })

                    except Exception as save_error:
                        results['errors'].append({'row': row_number, 'error': f"Save Error: {save_error}"})
                else:
                    results['errors'].append({'row': row_number, 'error': f"Validation failed: {serializer.errors}"})
                                
            except Exception as e:
                results['errors'].append({'row': index + 2, 'error': str(e)})
        
        # Update job statistics
        job.successful_records = results['successful_imports']
        job.failed_records = len(results['errors'])
        job.processed_records = results['processed_rows']
        job.save()
        
        job.update_progress(100, f"Import completed: {results['successful_imports']} students created")
        
        return results


class BulkImportStatusView(APIView):
    """
    View to check bulk import status and get available IDs for template
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        """Get available IDs for educational levels, grades, classes, academic years"""
        from schools.models import EducationalLevel, Grade, SchoolClass, AcademicYear
        
        try:
            # Get educational levels with their grades
            levels_data = []
            for level in EducationalLevel.objects.prefetch_related('grades').all():
                level_info = {
                    'id': level.id,
                    'name': level.name,
                    'grades': [{'id': g.id, 'name': g.name} for g in level.grades.all()]
                }
                levels_data.append(level_info)
            
            # Get current academic year
            current_year = AcademicYear.objects.filter(is_current=True).first()
            
            # Get all academic years
            academic_years = [
                {'id': ay.id, 'year': ay.year, 'is_current': ay.is_current}
                for ay in AcademicYear.objects.all().order_by('-year')
            ]
            
            # Get sample classes (limited to 10 for template reference)
            sample_classes = [
                {'id': sc.id, 'name': sc.name, 'grade_id': sc.grade_id}
                for sc in SchoolClass.objects.select_related('grade').all()[:10]
            ]
            
            return Response({
                'educational_levels': levels_data,
                'academic_years': academic_years,
                'current_academic_year': {
                    'id': current_year.id,
                    'year': current_year.year
                } if current_year else None,
                'sample_classes': sample_classes,
                'instructions': [
                    'Use the IDs from educational_levels and their grades',
                    'Class ID must match the selected grade',
                    f'Current academic year ID is: {current_year.id if current_year else "Not set"}',
                    'Contact administrator if you need specific class IDs'
                ]
            })
            
        except Exception as e:
            return Response(
                {'error': f'Failed to get import info: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class BulkImportProgressView(APIView):
    """
    View to check the progress of a bulk import job
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request, job_id):
        """Get progress status of a bulk import job"""
        try:
            # Find the job
            job = BulkImportJob.objects.get(job_id=job_id, created_by=request.user)
            
            response_data = {
                'job_id': str(job.job_id),
                'status': job.status,
                'progress': job.progress,
                'current_status': job.current_status or 'Processing...',
                'total_records': job.total_records,
                'processed_records': job.processed_records,
                'successful_records': job.successful_records,
                'failed_records': job.failed_records,
                'completed': job.is_completed,
                'created_at': job.created_at,
                'started_at': job.started_at,
                'completed_at': job.completed_at,
            }
            
            # Include error message if job failed
            if job.status == BulkImportJob.Status.FAILED:
                response_data['error'] = job.error_message
            
            # Include results if job completed successfully
            if job.status == BulkImportJob.Status.COMPLETED and job.results:
                response_data['results'] = job.results
            
            return Response(response_data, status=status.HTTP_200_OK)
            
        except BulkImportJob.DoesNotExist:
            return Response(
                {'error': 'Import job not found'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': f'Failed to get job status: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
